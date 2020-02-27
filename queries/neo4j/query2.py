from time import time

from neo4j import GraphDatabase

from openpyxl import Workbook
from openpyxl.styles import Alignment

NEO_LOGIN = "neo4j"
NEO_PASSWORD = "bitnami"

uri = "bolt://localhost:7687"
driver = GraphDatabase.driver(uri, auth=(NEO_LOGIN, NEO_PASSWORD))

start = time()
with driver.session() as session:
    data = session.run(
        "match (a1:Actor) - [:FILMED_IN] -> (f:Film) <- [:FILMED_IN] - (a2:Actor)\n"
        "where a1.actor_id < a2.actor_id\n"
        "return a1.actor_id as actor1_id, a2.actor_id as actor2_id, count(*) as count, "
        "a1.first_name as a1_first_name, a1.last_name as a1_last_name, "
        "a2.first_name as a2_first_name, a2.last_name as a2_last_name\n"
        "order by actor1_id, actor2_id"
    )
    q2 = [(x["actor1_id"], x["actor2_id"], x["count"],
           f"{x['a1_first_name']} {x['a1_last_name']}", f"{x['a2_first_name']} {x['a2_last_name']}") for x in data]

    query_executed = time() - start

    wb = Workbook()
    ws = wb.active
    m = -1
    for t in q2:
        actor1, actor2, count, name1, name2 = t
        ws.cell(row=actor1 + 1 + 1, column=actor2 + 1 + 1, value=count)
        ws.cell(row=actor2 + 1 + 1, column=actor1 + 1 + 1, value=count)

        ws.cell(row=actor1 + 1 + 1, column=1, value=name1)
        ws.cell(row=actor2 + 1 + 1, column=1, value=name2)

        ws.cell(row=1, column=actor1 + 1 + 1, value=name1).alignment = Alignment(text_rotation=85)
        ws.cell(row=1, column=actor2 + 1 + 1, value=name2).alignment = Alignment(text_rotation=85)

        m = max(m, actor1, actor2)

    for i in range(2, ws.max_row + 1):
        if i - 1 <= m:
            ws.cell(row=2, column=i + 1, value=i - 1)
            ws.cell(row=i + 1, column=2, value=i - 1)
        for j in range(3, ws.max_column + 1):
            cell = ws.cell(row=i, column=j)
            if cell.value is None:
                cell.value = 0

    wb.save("Query2.xlsx")
    print('See the result of query2 in the file "Query2.xlsx"')
print(f"Total time elapsed: {time() - start} sec\nQuery was executed in {query_executed}sec")
driver.close()
